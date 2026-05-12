from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
import time
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable
from urllib.parse import urljoin
import xml.etree.ElementTree as ET

import matplotlib.dates as mdates
import matplotlib.pyplot as plt
import requests
from openpyxl import Workbook
from matplotlib.ticker import FuncFormatter


BASE_URL = "https://sepe.es/HomeSepe/que-es-el-sepe/estadisticas/estadisticas-prestaciones/informe-prestaciones.html"
RAW_DIR = Path("data/raw")
PROCESSED_DIR = Path("data/processed")
FIGURES_DIR = Path("Gráficos")
FIGURE_WORKBOOKS_DIR = Path("data/figure_workbooks")
INTERACTIVE_DIR = Path("data/interactive")
MANIFEST_PATH = Path("data/manifest.json")

TARGET_SHEETS = {
    "BP-2.1a": ("total prestacion contributiva", "Ambos sexos", "age"),
    "BP-2.1b": ("total prestacion contributiva", "Hombres", "age"),
    "BP-2.1c": ("total prestacion contributiva", "Mujeres", "age"),
    "BP-3.1a": ("total subsidios de desempleo", "Ambos sexos", "age"),
    "BP-3.1b": ("total subsidios de desempleo", "Hombres", "age"),
    "BP-3.1c": ("total subsidios de desempleo", "Mujeres", "age"),
    "BP-3.5a": ("total subsidios de desempleo", "Ambos sexos", "subsidy_type"),
    "BP-3.5b": ("total subsidios de desempleo", "Hombres", "subsidy_type"),
    "BP-3.5c": ("total subsidios de desempleo", "Mujeres", "subsidy_type"),
    "TC-1.1a": ("tasa de cobertura", "Ambos sexos", "coverage"),
    "TC-1.1b": ("tasa de cobertura", "Hombres", "coverage"),
    "TC-1.1c": ("tasa de cobertura", "Mujeres", "coverage"),
}

MONTHS = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "setiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}

PROVINCE_TO_CCAA = {
    "Almería": "Andalucía", "Cádiz": "Andalucía", "Córdoba": "Andalucía", "Granada": "Andalucía",
    "Huelva": "Andalucía", "Jaén": "Andalucía", "Málaga": "Andalucía", "Sevilla": "Andalucía",
    "Huesca": "Aragón", "Teruel": "Aragón", "Zaragoza": "Aragón",
    "Asturias": "Asturias, Principado de", "Balears, Illes": "Balears, Illes",
    "Palmas, Las": "Canarias", "Santa Cruz de Tenerife": "Canarias",
    "Cantabria": "Cantabria", "Ávila": "Castilla y León", "Burgos": "Castilla y León",
    "León": "Castilla y León", "Palencia": "Castilla y León", "Salamanca": "Castilla y León",
    "Segovia": "Castilla y León", "Soria": "Castilla y León", "Valladolid": "Castilla y León",
    "Zamora": "Castilla y León", "Albacete": "Castilla-La Mancha", "Ciudad Real": "Castilla-La Mancha",
    "Cuenca": "Castilla-La Mancha", "Guadalajara": "Castilla-La Mancha", "Toledo": "Castilla-La Mancha",
    "Barcelona": "Cataluña", "Girona": "Cataluña", "Lleida": "Cataluña", "Tarragona": "Cataluña",
    "Alicante/Alacant": "Comunitat Valenciana", "Castellón/Castelló": "Comunitat Valenciana",
    "Valencia/València": "Comunitat Valenciana", "Badajoz": "Extremadura", "Cáceres": "Extremadura",
    "Coruña, A": "Galicia", "Lugo": "Galicia", "Ourense": "Galicia", "Pontevedra": "Galicia",
    "Madrid": "Madrid, Comunidad de", "Murcia": "Murcia, Región de", "Navarra": "Navarra, Comunidad Foral de",
    "Araba/Álava": "País Vasco", "Bizkaia": "País Vasco", "Gipuzkoa": "País Vasco",
    "Rioja, La": "Rioja, La", "Ceuta": "Ceuta", "Melilla": "Melilla",
}
CCAA_NAMES = set(PROVINCE_TO_CCAA.values())
GEO_ALIASES = {
    "A Coruña": "Coruña, A",
    "La Coruña": "Coruña, A",
    "Las Palmas": "Palmas, Las",
    "Comunidad Valenciana": "Comunitat Valenciana",
    "Navarra, Comunidad Foral": "Navarra, Comunidad Foral de",
    "Illes Balears": "Balears, Illes",
    "Baleares": "Balears, Illes",
    "La Rioja": "Rioja, La",
    "País Vasco": "País Vasco",
}

LONG_FIELDS = [
    "periodo", "año", "mes", "archivo_origen", "url_origen", "hoja_origen",
    "metrica", "variable", "variable_original", "sexo", "edad",
    "nivel_geografico", "provincia", "comunidad_autonoma", "valor",
]

WIDE_KEY_FIELDS = ["mes", "año", "sexo", "provincia", "edad", "comunidad autonoma", "nivel geografico"]
WIDE_VALUE_FIELDS = [
    "total prestacion contributiva",
    "total subsidios de desempleo",
    "subsidios de desempleo de mayores",
    "subsidio de desempleo por agotamiento de la prestacion contributiva",
    "subsidio de desempleo por no cotizacion suficiente",
    "subsidio de desempleo por no cotizacion suficiente - derecho de 3 a 5 meses",
    "subsidio de desempleo por no cotizacion suficiente - derecho de 6 meses",
    "subsidio de desempleo por no cotizacion suficiente - derecho de 21 meses",
    "subsidio de desempleo para emigrantes retornados",
    "subsidio de desempleo para liberados de prision",
    "subsidio de desempleo por revision de invalidez",
    "subsidio de desempleo para fijos discontinuos",
    "subsidio extraordinario por desempleo (SED)",
    "subsidio VVGS",
    "complemento de apoyo al empleo (CAE)",
    "tasa de cobertura",
]


@dataclass(frozen=True)
class RemoteFile:
    url: str
    year: int
    month: int
    filename: str


def norm_text(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\n", " ")).strip()


def sheet_key(name: str) -> str:
    return norm_text(name).rstrip(".")


def discover_files(url: str = BASE_URL) -> list[RemoteFile]:
    text = request("GET", url).text
    links = sorted(set(re.findall(r'href=["\']([^"\']+\.xlsx?[^"\']*)["\']', text, flags=re.I)))
    files = []
    for link in links:
        full_url = urljoin(url, link)
        match = re.search(r"(20\d{2})(0[1-9]|1[0-2])", full_url)
        if not match:
            continue
        year, month = int(match.group(1)), int(match.group(2))
        filename = f"Informe-{year}{month:02d}{Path(full_url.split('?')[0]).suffix.lower()}"
        files.append(RemoteFile(full_url, year, month, filename))
    return sorted(files, key=lambda f: (f.year, f.month, f.url))


def request(method: str, url: str, **kwargs) -> requests.Response:
    headers = kwargs.pop("headers", {})
    headers.setdefault("User-Agent", "Mozilla/5.0 sepe-prestaciones-pipeline/1.0")
    session = requests.Session()
    session.trust_env = False
    for attempt in range(4):
        try:
            response = session.request(method, url, headers=headers, timeout=60, **kwargs)
            response.raise_for_status()
            return response
        except requests.RequestException:
            if attempt == 3:
                raise
            time.sleep(2 ** attempt)
    raise RuntimeError("unreachable")


def load_manifest() -> dict:
    if MANIFEST_PATH.exists():
        return json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))
    return {"files": {}}


def save_manifest(manifest: dict) -> None:
    MANIFEST_PATH.parent.mkdir(parents=True, exist_ok=True)
    MANIFEST_PATH.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")


def sha256_bytes(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def download_files(files: list[RemoteFile], limit: int | None = None) -> list[Path]:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    manifest = load_manifest()
    paths = []
    selected = files[:limit] if limit else files
    for remote in selected:
        target = RAW_DIR / remote.filename
        entry = manifest["files"].get(remote.url, {})
        if target.exists() and entry.get("sha256"):
            head = try_head(remote.url)
            if head:
                same_size = str(target.stat().st_size) == head.get("content-length", "")
                same_etag = not head.get("etag") or head.get("etag") == entry.get("etag")
                same_modified = not head.get("last-modified") or head.get("last-modified") == entry.get("last_modified")
                if same_size and same_etag and same_modified:
                    paths.append(target)
                    continue
        response = request("GET", remote.url)
        digest = sha256_bytes(response.content)
        if target.exists() and entry.get("sha256") == digest:
            manifest["files"][remote.url] = {
                **entry,
                "year": remote.year,
                "month": remote.month,
                "filename": remote.filename,
                "local_path": str(target.as_posix()),
                "source_url": remote.url,
                "bytes": len(response.content),
                "etag": response.headers.get("ETag"),
                "last_modified": response.headers.get("Last-Modified"),
                "checked_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
            }
            paths.append(target)
            continue
        target.write_bytes(response.content)
        manifest["files"][remote.url] = {
            "year": remote.year,
            "month": remote.month,
            "filename": remote.filename,
            "local_path": str(target.as_posix()),
            "source_url": remote.url,
            "sha256": digest,
            "bytes": len(response.content),
            "etag": response.headers.get("ETag"),
            "last_modified": response.headers.get("Last-Modified"),
            "downloaded_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        }
        paths.append(target)
    save_manifest(manifest)
    return paths


def try_head(url: str) -> dict[str, str] | None:
    try:
        response = request("HEAD", url)
        return {k.lower(): v for k, v in response.headers.items()}
    except requests.RequestException:
        return None


MAX_SCAN_COLS = 80
MAX_SCAN_ROWS = 220


def iter_data_rows(rows: dict[int, list], start_row: int) -> Iterable[tuple[int, list]]:
    blank_count = 0
    for row in range(start_row, MAX_SCAN_ROWS + 1):
        values = rows.get(row, [])
        if not any(v not in (None, "") for v in values):
            blank_count += 1
            if blank_count >= 3:
                break
            continue
        blank_count = 0
        yield row, values


def geography(name: str) -> tuple[str, str, str]:
    name = norm_text(name)
    name = GEO_ALIASES.get(name, name)
    total_names = {"Total Nacional", "TOTAL NACIONAL", "Total", "TOTAL", "España"}
    if name in total_names or name.upper() == "TOTAL NACIONAL":
        return "espana", "España", "España"
    if name in PROVINCE_TO_CCAA:
        return "provincia", name, PROVINCE_TO_CCAA[name]
    if name in CCAA_NAMES:
        return "comunidad_autonoma", "Todas las provincias", name
    return "desconocido", name, "Desconocida"


def clean_variable(label: str) -> str:
    label = norm_text(label)
    label = label.replace("Mayores de 55 años", "Mayores de 52/55 años")
    label = label.replace("Mayores de 52 años", "Mayores de 52/55 años")
    return label


def variable_name(metric: str, original: str, kind: str) -> str:
    label = clean_variable(original).lower()
    if kind in {"age", "coverage"}:
        return metric
    if label == "total":
        return "total subsidios de desempleo"
    if "mayores de 52/55" in label:
        return "subsidios de desempleo de mayores"
    if "agotamiento prestación contributiva" in label:
        return "subsidio de desempleo por agotamiento de la prestacion contributiva"
    if "periodo cotizado insuficiente" in label:
        suffix = ""
        if "derecho de" in label:
            suffix = " - " + label.split(" - ", 1)[-1]
        return "subsidio de desempleo por no cotizacion suficiente" + suffix
    if "emigrantes retornados" in label:
        return "subsidio de desempleo para emigrantes retornados"
    if "liberados de prisión" in label:
        return "subsidio de desempleo para liberados de prision"
    if "fijos discontinuos" in label:
        return "subsidio de desempleo para fijos discontinuos"
    if "plenamente capaces" in label or "inválidos parciales" in label:
        return "subsidio de desempleo por revision de invalidez"
    if "subsidio extraordinario por desempleo" in label:
        return "subsidio extraordinario por desempleo (SED)"
    if "subsidio vvgs" in label:
        return "subsidio VVGS"
    if "complemento de apoyo al empleo" in label:
        return "complemento de apoyo al empleo (CAE)"
    return clean_variable(original)


def parse_workbook(path: Path, source_url: str | None = None) -> list[dict]:
    match = re.search(r"(20\d{2})(0[1-9]|1[0-2])", path.name)
    if not match:
        raise ValueError(f"Cannot infer period from {path}")
    year, month = int(match.group(1)), int(match.group(2))
    period = f"{year}-{month:02d}"
    workbook = read_xlsx(path)
    sheets = {sheet_key(name): name for name in workbook}
    records = []
    for target, (metric, sex, kind) in TARGET_SHEETS.items():
        sheet_name = sheets.get(target)
        if not sheet_name:
            continue
        rows = workbook[sheet_name]
        if kind == "coverage":
            records.extend(parse_coverage_sheet(rows, path, source_url, period, year, month, target, metric, sex))
        else:
            records.extend(parse_table_sheet(rows, path, source_url, period, year, month, target, metric, sex, kind))
    return records


def parse_table_sheet(rows: dict[int, list], path: Path, source_url: str | None, period: str, year: int, month: int,
                      sheet: str, metric: str, sex: str, kind: str) -> list[dict]:
    header_row = find_row_containing(rows, "Provincias y CC.AA.")
    if not header_row:
        return []
    top = rows.get(header_row - 1, [])
    bottom = rows.get(header_row, [])
    headers = build_headers(top, bottom, kind)
    records = []
    for _, values in iter_data_rows(rows, header_row + 1):
        if len(values) < 3 or not values[1]:
            continue
        geo_level, province, ccaa = geography(values[1])
        for idx, variable in headers.items():
            if idx >= len(values):
                continue
            value = parse_number(values[idx])
            if value is None:
                continue
            original = variable
            age = "Todas las edades"
            if kind == "age":
                age = "Todas las edades" if variable == "TOTAL" else clean_variable(variable)
                variable = metric
            else:
                variable = variable_name(metric, variable, kind)
            records.append(record(period, year, month, path, source_url, sheet, metric, variable,
                                  original, sex, age, geo_level, province, ccaa, value))
    return records


def build_headers(top: list, bottom: list, kind: str) -> dict[int, str]:
    headers = {}
    current_group = None
    max_len = max(len(top), len(bottom))
    for idx in range(2, max_len):
        top_label = norm_text(top[idx] if idx < len(top) else "")
        bottom_label = norm_text(bottom[idx] if idx < len(bottom) else "")
        if top_label and top_label not in {"Tramos de edad"}:
            current_group = top_label
        if kind == "age":
            label = "TOTAL" if top_label == "TOTAL" else bottom_label
        else:
            label = top_label
            if current_group and bottom_label and top_label in {"", current_group}:
                label = f"{current_group} - {bottom_label}"
        if label:
            headers[idx] = label
    return headers


def parse_coverage_sheet(rows: dict[int, list], path: Path, source_url: str | None, period: str, year: int, month: int,
                         sheet: str, metric: str, sex: str) -> list[dict]:
    header_row = find_row_containing(rows, "Provincias y CC.AA.")
    if not header_row:
        return []
    headers = rows.get(header_row, [])
    month_col = None
    for idx, label in enumerate(headers):
        if MONTHS.get(norm_text(label).lower()) == month:
            month_col = idx
            break
    if month_col is None:
        return []
    records = []
    for _, values in iter_data_rows(rows, header_row + 1):
        if len(values) <= month_col or not values[1]:
            continue
        value = parse_number(values[month_col])
        if value is None:
            continue
        geo_level, province, ccaa = geography(values[1])
        records.append(record(period, year, month, path, source_url, sheet, metric, metric,
                              norm_text(headers[month_col]), sex, "Todas las edades", geo_level, province, ccaa, value))
    return records


def find_row_containing(rows: dict[int, list], needle: str) -> int | None:
    for row in range(1, 81):
        if any(norm_text(value) == needle for value in rows.get(row, [])[:20]):
            return row
    return None


def read_xlsx(path: Path) -> dict[str, dict[int, list]]:
    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "pkgrel": "http://schemas.openxmlformats.org/package/2006/relationships",
    }
    with zipfile.ZipFile(path) as zf:
        shared = read_shared_strings(zf, ns)
        rels = read_workbook_rels(zf, ns)
        root = ET.fromstring(zf.read("xl/workbook.xml"))
        out = {}
        for sheet in root.findall(".//main:sheet", ns):
            name = sheet.attrib["name"]
            if sheet_key(name) not in TARGET_SHEETS:
                continue
            rel_id = sheet.attrib.get(f"{{{ns['rel']}}}id")
            target = rels.get(rel_id)
            if not target:
                continue
            sheet_path = "xl/" + target.lstrip("/")
            out[name] = read_sheet_rows(zf, sheet_path, shared, ns)
        return out


def read_shared_strings(zf: zipfile.ZipFile, ns: dict[str, str]) -> list[str]:
    if "xl/sharedStrings.xml" not in zf.namelist():
        return []
    root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
    strings = []
    for si in root.findall("main:si", ns):
        parts = [t.text or "" for t in si.findall(".//main:t", ns)]
        strings.append("".join(parts))
    return strings


def read_workbook_rels(zf: zipfile.ZipFile, ns: dict[str, str]) -> dict[str, str]:
    root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rels = {}
    for rel in root.findall("pkgrel:Relationship", ns):
        rels[rel.attrib["Id"]] = rel.attrib["Target"]
    return rels


def read_sheet_rows(zf: zipfile.ZipFile, sheet_path: str, shared: list[str], ns: dict[str, str]) -> dict[int, list]:
    rows = {}
    for event, elem in ET.iterparse(zf.open(sheet_path), events=("end",)):
        if elem.tag.rsplit("}", 1)[-1] != "row":
            continue
        row_idx = int(elem.attrib.get("r", "0"))
        if row_idx > MAX_SCAN_ROWS:
            elem.clear()
            break
        values = []
        for cell in elem:
            if cell.tag.rsplit("}", 1)[-1] != "c":
                continue
            col_idx = column_index(cell.attrib.get("r", "A1"))
            if col_idx > MAX_SCAN_COLS:
                continue
            while len(values) < col_idx:
                values.append(None)
            values[col_idx - 1] = cell_value(cell, shared)
        while values and values[-1] in (None, ""):
            values.pop()
        if values:
            rows[row_idx] = values
        elem.clear()
    return rows


def column_index(ref: str) -> int:
    letters = re.match(r"([A-Z]+)", ref or "A").group(1)
    value = 0
    for char in letters:
        value = value * 26 + ord(char) - 64
    return value


def cell_value(cell, shared: list[str]):
    ctype = cell.attrib.get("t")
    value = cell.find("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v")
    if ctype == "inlineStr":
        return "".join(t.text or "" for t in cell.findall(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t"))
    if value is None or value.text is None:
        return None
    if ctype == "s":
        return shared[int(value.text)]
    if ctype == "str":
        return value.text
    try:
        number = float(value.text)
        return int(number) if number.is_integer() else number
    except ValueError:
        return value.text


def parse_number(value):
    if value in (None, "", "-"):
        return None
    if isinstance(value, (int, float)):
        return value
    text = norm_text(value).replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def record(period, year, month, path, source_url, sheet, metric, variable, original, sex, age,
           geo_level, province, ccaa, value) -> dict:
    return {
        "periodo": period,
        "año": year,
        "mes": month,
        "archivo_origen": path.name,
        "url_origen": source_url or "",
        "hoja_origen": sheet,
        "metrica": metric,
        "variable": variable,
        "variable_original": original,
        "sexo": sex,
        "edad": age,
        "nivel geografico": geo_level,
        "provincia": province,
        "comunidad autonoma": ccaa,
        "valor": value,
    }


def export_records(records: list[dict]) -> None:
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    write_wide(make_wide(records, all_ages_only=False), PROCESSED_DIR / "sepe_prestaciones_wide")
    generate_figures(PROCESSED_DIR / "sepe_prestaciones_wide.csv")


def write_wide(wide_rows: list[dict], stem: Path) -> None:
    if not wide_rows:
        return
    value_set = {key for row in wide_rows for key in row if key not in WIDE_KEY_FIELDS}
    value_fields = [field for field in WIDE_VALUE_FIELDS if field in value_set]
    fields = WIDE_KEY_FIELDS + value_fields
    csv_path = stem.with_suffix(".csv")
    with csv_path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(wide_rows)
    write_xlsx(wide_rows, stem.with_suffix(".xlsx"), fields)


def generate_figures(csv_path: Path) -> list[Path]:
    if not csv_path.exists():
        return []

    rows = read_national_all_ages(csv_path)
    if not rows:
        return []

    FIGURES_DIR.mkdir(parents=True, exist_ok=True)
    plt.style.use("seaborn-v0_8-whitegrid")
    plt.rcParams.update({
        "axes.edgecolor": "#404040",
        "axes.labelcolor": "#404040",
        "axes.titlecolor": "#404040",
        "axes.spines.right": False,
        "axes.spines.top": False,
        "grid.color": "#D9D9D9",
        "grid.linewidth": 0.6,
        "figure.facecolor": "white",
        "font.family": "Century Gothic",
        "font.size": 9,
        "savefig.dpi": 180,
        "svg.fonttype": "none",
    })

    static_outputs = [
        plot_beneficiaries_and_coverage(rows),
        plot_benefit_mix(rows),
        plot_coverage_vs_beneficiaries_index(rows),
        plot_age_profile(csv_path),
        plot_gender_share(csv_path),
        plot_regional_coverage_latest(csv_path),
        plot_regional_dispersion(csv_path),
    ]
    png_outputs = [
        path.with_suffix(".png")
        for path in static_outputs
        if path and path.suffix.lower() == ".svg" and path.with_suffix(".png").exists()
    ]
    outputs = static_outputs + png_outputs
    outputs.extend(generate_interactive_graphs(csv_path))
    return [path for path in outputs if path]


def read_national_all_ages(csv_path: Path) -> list[dict]:
    wanted = {
        "total prestacion contributiva",
        "total subsidios de desempleo",
        "tasa de cobertura",
    }
    rows = {}
    with csv_path.open(newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            if row.get("sexo") != "Ambos sexos":
                continue
            if row.get("edad") != "Todas las edades":
                continue
            if row.get("nivel geografico") != "espana":
                continue
            period = date(int(row["año"]), int(row["mes"]), 1)
            current = rows.setdefault(period, {"period": period})
            for field in wanted:
                value = parse_csv_number(row.get(field))
                if value is not None:
                    current[field] = value
    return [rows[key] for key in sorted(rows)]


def parse_csv_number(value):
    if value in (None, ""):
        return None
    try:
        return float(value)
    except ValueError:
        return parse_number(value)


def plot_beneficiaries_and_coverage(rows: list[dict]) -> Path:
    periods = [row["period"] for row in rows]
    contributiva = [row.get("total prestacion contributiva") for row in rows]
    subsidios = [row.get("total subsidios de desempleo") for row in rows]
    total = [safe_sum(a, b) for a, b in zip(contributiva, subsidios)]
    coverage = [row.get("tasa de cobertura") for row in rows]

    fig, ax = plt.subplots(figsize=(6, 3))
    ax.plot(periods, total, color="#83082A", linewidth=2, label="Total beneficiarios")
    ax.plot(periods, contributiva, color="#D00D43", linewidth=1.5, label="Prestación contributiva")
    ax.plot(periods, subsidios, color="#E397A0", linewidth=1.5, label="Subsidios de desempleo")
    ax.yaxis.set_major_formatter(FuncFormatter(format_thousands))
    ax.xaxis.set_major_locator(mdates.YearLocator())
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
    ax.grid(axis="y")
    ax.grid(axis="x", visible=False)
    ax.tick_params(direction="out", colors="#404040")

    ax2 = ax.twinx()
    ax2.spines["top"].set_visible(False)
    ax2.plot(periods, coverage, color="#404040", linewidth=1.5, linestyle="--", label="Tasa de cobertura")
    ax2.tick_params(direction="out", colors="#404040")

    lines, labels = ax.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax.legend(lines + lines2, labels + labels2, loc="upper left", frameon=False, fontsize=8)
    write_chart_workbook(
        "evolucion_beneficiarios_tasa_cobertura.xlsx",
        "Evolución de la protección por desempleo en España",
        ["Periodo", "Total beneficiarios", "Prestación contributiva", "Subsidios de desempleo", "Tasa de cobertura"],
        zip(periods, total, contributiva, subsidios, coverage),
        "Fuente: AIReF a partir de SEPE.",
        "Nota: Beneficiarios en personas y tasa de cobertura en porcentaje.",
    )
    return save_figure(fig, "evolucion_beneficiarios_tasa_cobertura.svg")


def plot_benefit_mix(rows: list[dict]) -> Path:
    periods = [row["period"] for row in rows]
    contributiva = [row.get("total prestacion contributiva") or 0 for row in rows]
    subsidios = [row.get("total subsidios de desempleo") or 0 for row in rows]

    fig, ax = plt.subplots(figsize=(6, 3))
    ax.stackplot(
        periods,
        contributiva,
        subsidios,
        labels=["Prestación contributiva", "Subsidios de desempleo"],
        colors=["#83082A", "#E397A0"],
        alpha=0.92,
    )
    ax.yaxis.set_major_formatter(FuncFormatter(format_thousands))
    ax.xaxis.set_major_locator(mdates.YearLocator())
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
    ax.grid(axis="y")
    ax.grid(axis="x", visible=False)
    ax.tick_params(direction="out", colors="#404040")
    ax.legend(loc="upper left", frameon=False, fontsize=8)
    write_chart_workbook(
        "composicion_beneficiarios_prestaciones.xlsx",
        "Composición de beneficiarios por tipo de prestación",
        ["Periodo", "Prestación contributiva", "Subsidios de desempleo"],
        zip(periods, contributiva, subsidios),
        "Fuente: AIReF a partir de SEPE.",
        "Nota: Beneficiarios en personas.",
    )
    return save_figure(fig, "composicion_beneficiarios_prestaciones.svg")


def plot_coverage_vs_beneficiaries_index(rows: list[dict]) -> Path:
    indexed = [row for row in rows if row.get("tasa de cobertura") is not None]
    if not indexed:
        return Path()
    periods = [row["period"] for row in indexed]
    total = [safe_sum(row.get("total prestacion contributiva"), row.get("total subsidios de desempleo")) for row in indexed]
    coverage = [row.get("tasa de cobertura") for row in indexed]
    base_total = next((value for value in total if value), None)
    base_coverage = next((value for value in coverage if value), None)
    if not base_total or not base_coverage:
        return Path()

    fig, ax = plt.subplots(figsize=(6, 3))
    ax.axhline(100, color="#8a94a6", linewidth=1, linestyle=":")
    beneficiaries_index = [value / base_total * 100 if value else None for value in total]
    coverage_index = [value / base_coverage * 100 if value else None for value in coverage]
    ax.plot(periods, beneficiaries_index, color="#83082A", linewidth=2, label="Beneficiarios")
    ax.plot(periods, coverage_index, color="#404040", linewidth=1.5, label="Tasa de cobertura")
    ax.xaxis.set_major_locator(mdates.YearLocator())
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
    ax.grid(axis="y")
    ax.grid(axis="x", visible=False)
    ax.tick_params(direction="out", colors="#404040")
    ax.legend(loc="upper left", frameon=False, fontsize=8)
    write_chart_workbook(
        "indice_beneficiarios_tasa_cobertura.xlsx",
        "Beneficiarios y cobertura, índice base primer mes = 100",
        ["Periodo", "Beneficiarios", "Tasa de cobertura"],
        zip(periods, beneficiaries_index, coverage_index),
        "Fuente: AIReF a partir de SEPE.",
        "Nota: Índice base primer mes = 100.",
    )
    return save_figure(fig, "indice_beneficiarios_tasa_cobertura.svg")


def plot_age_profile(csv_path: Path) -> Path:
    rows = read_age_profile(csv_path)
    if not rows:
        return Path()
    ages = [row["age"] for row in rows]
    contributiva = [row["contributiva"] for row in rows]
    subsidios = [row["subsidios"] for row in rows]

    fig, ax = plt.subplots(figsize=(6, 3))
    x = range(len(rows))
    ax.bar(x, contributiva, color="#83082A", label="Prestación contributiva")
    ax.bar(x, subsidios, bottom=contributiva, color="#E397A0", label="Subsidios de desempleo")
    ax.set_xticks(list(x))
    ax.set_xticklabels(ages, rotation=35, ha="right")
    ax.yaxis.set_major_formatter(FuncFormatter(format_thousands))
    ax.grid(axis="y")
    ax.grid(axis="x", visible=False)
    ax.tick_params(direction="out", colors="#404040")
    ax.legend(loc="upper left", frameon=False, fontsize=8)
    write_chart_workbook(
        "perfil_edad_beneficiarios.xlsx",
        f"Perfil por edad de los beneficiarios ({rows[0]['period_label']})",
        ["Edad", "Prestación contributiva", "Subsidios de desempleo", "Total beneficiarios"],
        ((row["age"], row["contributiva"], row["subsidios"], row["total"]) for row in rows),
        "Fuente: AIReF a partir de SEPE.",
        "Nota: Beneficiarios en personas. Datos nacionales para ambos sexos.",
    )
    return save_figure(fig, "perfil_edad_beneficiarios.svg")


def plot_gender_share(csv_path: Path) -> Path:
    rows = read_gender_share(csv_path)
    if not rows:
        return Path()
    periods = [row["period"] for row in rows]
    women_share = [row["women_share"] for row in rows]
    men_share = [100 - value for value in women_share]

    fig, ax = plt.subplots(figsize=(6, 3))
    ax.plot(periods, women_share, color="#83082A", linewidth=2, label="Mujeres")
    ax.plot(periods, men_share, color="#404040", linewidth=1.5, label="Hombres")
    ax.axhline(50, color="#D9D9D9", linewidth=1, linestyle=":")
    ax.xaxis.set_major_locator(mdates.YearLocator())
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
    ax.yaxis.set_major_formatter(FuncFormatter(format_percent_no_symbol))
    ax.grid(axis="y")
    ax.grid(axis="x", visible=False)
    ax.tick_params(direction="out", colors="#404040")
    ax.legend(loc="upper left", frameon=False, fontsize=8)
    write_chart_workbook(
        "peso_beneficiarias_por_sexo.xlsx",
        "Peso de mujeres y hombres en el total de beneficiarios",
        ["Periodo", "Mujeres", "Hombres"],
        zip(periods, women_share, men_share),
        "Fuente: AIReF a partir de SEPE.",
        "Nota: Porcentaje sobre beneficiarios con prestación contributiva o subsidio de desempleo.",
    )
    return save_figure(fig, "peso_beneficiarias_por_sexo.svg")


def plot_regional_coverage_latest(csv_path: Path) -> Path:
    rows = read_regional_latest(csv_path)
    rows = [row for row in rows if row.get("coverage") is not None]
    if not rows:
        return Path()
    spain_coverage = read_latest_national_coverage(csv_path)
    rows = sorted(rows, key=lambda row: row["coverage"])
    regions = [short_region_name(row["region"]) for row in rows]
    values = [row["coverage"] for row in rows]
    colors = ["#83082A" if value == max(values) else "#E397A0" for value in values]

    fig, ax = plt.subplots(figsize=(6, 3))
    y = range(len(rows))
    ax.barh(y, values, color=colors)
    ax.set_yticks(list(y))
    ax.set_yticklabels(regions)
    ax.xaxis.set_major_formatter(FuncFormatter(format_percent_no_symbol))
    ax.grid(axis="x")
    ax.grid(axis="y", visible=False)
    ax.tick_params(direction="out", colors="#404040")
    if spain_coverage is not None:
        ax.axvline(spain_coverage, color="#404040", linewidth=2, label="España")
        ax.legend(loc="lower right", frameon=False, fontsize=8)
    write_chart_workbook(
        "tasa_cobertura_ccaa_ultimo_periodo.xlsx",
        f"Tasa de cobertura por comunidad autónoma ({rows[0]['period_label']})",
        ["Comunidad autónoma", "Tasa de cobertura", "España"],
        ((row["region"], row["coverage"], spain_coverage) for row in rows),
        "Fuente: AIReF a partir de SEPE.",
        "Nota: Tasa de cobertura en porcentaje. Se muestran las comunidades autónomas identificadas en la base procesada.",
    )
    return save_figure(fig, "tasa_cobertura_ccaa_ultimo_periodo.svg")


def generate_interactive_graphs(csv_path: Path) -> list[Path]:
    INTERACTIVE_DIR.mkdir(parents=True, exist_ok=True)
    return [
        write_interactive_regional_coverage(read_regional_coverage_by_year(csv_path)),
        write_interactive_age_profile(read_age_profile_by_year(csv_path)),
        write_interactive_regional_subsidy_share(read_regional_subsidy_share_by_year(csv_path)),
        write_interactive_age_gender_share(read_age_gender_share_by_year(csv_path)),
        write_interactive_coverage_heatmap(read_regional_coverage_by_year(csv_path)),
    ]


def write_interactive_regional_coverage(data: dict[int, dict]) -> Path:
    path = INTERACTIVE_DIR / "tasa_cobertura_ccaa_ultimo_periodo.html"
    path.write_text(interactive_html(
        "Tasa de cobertura por comunidad autónoma",
        "tasa_cobertura_ccaa",
        data,
        REGIONAL_COVERAGE_JS,
    ), encoding="utf-8")
    return path


def write_interactive_age_profile(data: dict[int, list[dict]]) -> Path:
    path = INTERACTIVE_DIR / "perfil_edad_beneficiarios.html"
    path.write_text(interactive_html(
        "Perfil por edad de los beneficiarios",
        "perfil_edad",
        data,
        AGE_PROFILE_JS,
    ), encoding="utf-8")
    return path


def write_interactive_regional_subsidy_share(data: dict[int, dict]) -> Path:
    path = INTERACTIVE_DIR / "peso_subsidios_ccaa.html"
    path.write_text(interactive_html(
        "Peso de subsidios en el total de beneficiarios",
        "peso_subsidios_ccaa",
        data,
        REGIONAL_SUBSIDY_SHARE_JS,
    ), encoding="utf-8")
    return path


def write_interactive_age_gender_share(data: dict[int, list[dict]]) -> Path:
    path = INTERACTIVE_DIR / "peso_mujeres_por_edad.html"
    path.write_text(interactive_html(
        "Peso de mujeres por tramo de edad",
        "peso_mujeres_edad",
        data,
        AGE_GENDER_SHARE_JS,
    ), encoding="utf-8")
    return path


def write_interactive_coverage_heatmap(data: dict[int, dict]) -> Path:
    path = INTERACTIVE_DIR / "mapa_calor_tasa_cobertura_ccaa.html"
    path.write_text(interactive_html(
        "Tasa de cobertura por comunidad autónoma y año",
        "mapa_calor_cobertura",
        data,
        COVERAGE_HEATMAP_JS,
    ), encoding="utf-8")
    return path


def interactive_html(title: str, chart_id: str, data, renderer: str) -> str:
    payload = json.dumps(data, ensure_ascii=False)
    return f"""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{title}</title>
  <style>
    :root {{
      --airef-burgundy: #83082A;
      --airef-crimson: #D00D43;
      --airef-rose: #E397A0;
      --airef-text: #404040;
      --airef-grid: #D9D9D9;
      --panel: #ffffff;
      --page: #f7f5f5;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      background: linear-gradient(180deg, #ffffff 0%, var(--page) 100%);
      color: var(--airef-text);
      font-family: "Century Gothic", "Aptos", sans-serif;
    }}
    main {{
      max-width: 980px;
      margin: 0 auto;
      padding: 28px 22px 36px;
    }}
    .panel {{
      background: var(--panel);
      border: 1px solid #eee5e7;
      border-radius: 8px;
      box-shadow: 0 18px 45px rgba(64, 64, 64, 0.08);
      padding: 18px 18px 12px;
    }}
    .topline {{
      display: flex;
      align-items: end;
      justify-content: space-between;
      gap: 18px;
      margin-bottom: 14px;
    }}
    h1 {{
      margin: 0;
      color: var(--airef-text);
      font-size: 18px;
      font-weight: 700;
      letter-spacing: 0;
    }}
    .kpi {{
      color: var(--airef-burgundy);
      font-size: 30px;
      font-weight: 700;
      line-height: 1;
      white-space: nowrap;
    }}
    .controls {{
      display: grid;
      grid-template-columns: 1fr auto;
      align-items: center;
      gap: 14px;
      margin: 8px 0 16px;
    }}
    input[type="range"] {{
      width: 100%;
      accent-color: var(--airef-burgundy);
    }}
    .year-pill {{
      min-width: 72px;
      border: 1px solid #ead5db;
      border-radius: 999px;
      padding: 6px 12px;
      text-align: center;
      font-weight: 700;
      color: var(--airef-burgundy);
      background: #fff7f9;
    }}
    svg {{
      width: 100%;
      height: auto;
      display: block;
      overflow: visible;
    }}
    .axis text, .legend, .note {{
      fill: var(--airef-text);
      font-size: 12px;
    }}
    .grid line {{
      stroke: var(--airef-grid);
      stroke-width: 1;
    }}
    .tooltip {{
      position: fixed;
      pointer-events: none;
      transform: translate(12px, -12px);
      background: #ffffff;
      border: 1px solid #ead5db;
      border-radius: 6px;
      padding: 8px 10px;
      color: var(--airef-text);
      box-shadow: 0 12px 30px rgba(64, 64, 64, 0.14);
      font-size: 12px;
      opacity: 0;
      transition: opacity 120ms ease;
      z-index: 10;
    }}
    .source {{
      margin: 10px 0 0;
      font-size: 12px;
      color: var(--airef-text);
    }}
    @media (max-width: 680px) {{
      main {{ padding: 18px 12px 28px; }}
      .topline {{ align-items: start; flex-direction: column; }}
      .kpi {{ font-size: 24px; }}
    }}
  </style>
</head>
<body>
  <main>
    <section class="panel">
      <div class="topline">
        <h1>{title}</h1>
        <div class="kpi" id="headline"></div>
      </div>
      <div class="controls">
        <input id="yearSlider" type="range">
        <div class="year-pill" id="yearLabel"></div>
      </div>
      <svg id="{chart_id}" viewBox="0 0 920 430" role="img" aria-label="{title}"></svg>
      <p class="source">Fuente: AIReF a partir de SEPE.</p>
    </section>
  </main>
  <div class="tooltip" id="tooltip"></div>
  <script>
    const chartData = {payload};
{renderer}
  </script>
</body>
</html>
"""


REGIONAL_COVERAGE_JS = r"""
    const years = Object.keys(chartData).map(Number).sort((a, b) => a - b);
    const slider = document.getElementById("yearSlider");
    const yearLabel = document.getElementById("yearLabel");
    const headline = document.getElementById("headline");
    const svg = document.getElementById("tasa_cobertura_ccaa");
    const tooltip = document.getElementById("tooltip");
    slider.min = 0;
    slider.max = years.length - 1;
    slider.value = years.length - 1;
    slider.addEventListener("input", () => draw(years[Number(slider.value)]));
    const fixedMax = Math.ceil(Math.max(...years.flatMap(year => [
      chartData[year].spain,
      ...chartData[year].regions.map(row => row.coverage)
    ])) / 20) * 20;

    function el(name, attrs = {}, parent = svg) {
      const node = document.createElementNS("http://www.w3.org/2000/svg", name);
      for (const [key, value] of Object.entries(attrs)) node.setAttribute(key, value);
      parent.appendChild(node);
      return node;
    }
    function clear() { while (svg.firstChild) svg.removeChild(svg.firstChild); }
    function fmt(value) { return value.toLocaleString("es-ES", { maximumFractionDigits: 1 }); }
    function showTip(event, html) {
      tooltip.innerHTML = html;
      tooltip.style.left = `${event.clientX}px`;
      tooltip.style.top = `${event.clientY}px`;
      tooltip.style.opacity = 1;
    }
    function hideTip() { tooltip.style.opacity = 0; }
    function draw(year) {
      clear();
      yearLabel.textContent = year;
      const payload = chartData[year];
      const rows = [...payload.regions].sort((a, b) => a.coverage - b.coverage);
      const spain = payload.spain;
      headline.textContent = `${fmt(spain)} España`;
      const margin = { top: 18, right: 36, bottom: 42, left: 190 };
      const width = 920 - margin.left - margin.right;
      const height = 430 - margin.top - margin.bottom;
      const maxValue = fixedMax;
      const step = height / rows.length;
      const x = value => margin.left + value / maxValue * width;
      const y = index => margin.top + index * step + step * 0.16;

      for (let tick = 0; tick <= maxValue; tick += 20) {
        const tx = x(tick);
        el("line", { x1: tx, y1: margin.top, x2: tx, y2: margin.top + height, stroke: "#D9D9D9" });
        el("text", { x: tx, y: margin.top + height + 22, "text-anchor": "middle", fill: "#404040", "font-size": 12 }).textContent = tick;
      }

      const sx = x(spain);
      el("line", { x1: sx, y1: margin.top - 4, x2: sx, y2: margin.top + height, stroke: "#404040", "stroke-width": 3 });
      el("text", { x: sx + 6, y: margin.top + 12, fill: "#404040", "font-size": 12, "font-weight": 700 }).textContent = "España";

      rows.forEach((row, i) => {
        const barWidth = Math.max(0, x(row.coverage) - margin.left);
        el("text", { x: margin.left - 10, y: y(i) + step * 0.35, "text-anchor": "end", fill: "#404040", "font-size": 12 }).textContent = row.short;
        const rect = el("rect", {
          x: margin.left,
          y: y(i),
          width: barWidth,
          height: step * 0.58,
          fill: row.coverage >= spain ? "#83082A" : "#E397A0",
          rx: 2
        });
        rect.addEventListener("mousemove", event => showTip(event, `<strong>${row.region}</strong><br>${fmt(row.coverage)}<br>España: ${fmt(spain)}`));
        rect.addEventListener("mouseleave", hideTip);
        el("text", { x: x(row.coverage) + 6, y: y(i) + step * 0.38, fill: "#404040", "font-size": 11 }).textContent = fmt(row.coverage);
      });
    }
    draw(years[years.length - 1]);
"""


AGE_PROFILE_JS = r"""
    const years = Object.keys(chartData).map(Number).sort((a, b) => a - b);
    const slider = document.getElementById("yearSlider");
    const yearLabel = document.getElementById("yearLabel");
    const headline = document.getElementById("headline");
    const svg = document.getElementById("perfil_edad");
    const tooltip = document.getElementById("tooltip");
    slider.min = 0;
    slider.max = years.length - 1;
    slider.value = years.length - 1;
    slider.addEventListener("input", () => draw(years[Number(slider.value)]));
    const fixedMax = Math.ceil(Math.max(...years.flatMap(year => chartData[year].map(row => row.total))) / 100000) * 100000;

    function el(name, attrs = {}, parent = svg) {
      const node = document.createElementNS("http://www.w3.org/2000/svg", name);
      for (const [key, value] of Object.entries(attrs)) node.setAttribute(key, value);
      parent.appendChild(node);
      return node;
    }
    function clear() { while (svg.firstChild) svg.removeChild(svg.firstChild); }
    function fmt(value) { return value.toLocaleString("es-ES", { maximumFractionDigits: 0 }); }
    function fmtShort(value) {
      return value >= 1000000 ? `${(value / 1000000).toLocaleString("es-ES", { maximumFractionDigits: 1 })}M` : fmt(value);
    }
    function showTip(event, html) {
      tooltip.innerHTML = html;
      tooltip.style.left = `${event.clientX}px`;
      tooltip.style.top = `${event.clientY}px`;
      tooltip.style.opacity = 1;
    }
    function hideTip() { tooltip.style.opacity = 0; }
    function draw(year) {
      clear();
      yearLabel.textContent = year;
      const rows = chartData[year];
      const total = rows.reduce((sum, row) => sum + row.total, 0);
      headline.textContent = `${fmtShort(total)} total`;
      const margin = { top: 20, right: 26, bottom: 86, left: 72 };
      const width = 920 - margin.left - margin.right;
      const height = 430 - margin.top - margin.bottom;
      const maxValue = fixedMax;
      const band = width / rows.length;
      const x = index => margin.left + index * band + band * 0.18;
      const barWidth = band * 0.64;
      const y = value => margin.top + height - value / maxValue * height;

      for (let tick = 0; tick <= maxValue; tick += 100000) {
        const ty = y(tick);
        el("line", { x1: margin.left, y1: ty, x2: margin.left + width, y2: ty, stroke: "#D9D9D9" });
        el("text", { x: margin.left - 10, y: ty + 4, "text-anchor": "end", fill: "#404040", "font-size": 12 }).textContent = fmtShort(tick);
      }

      rows.forEach((row, i) => {
        const bx = x(i);
        const yContrib = y(row.contributiva);
        const yTotal = y(row.total);
        const contribRect = el("rect", { x: bx, y: yContrib, width: barWidth, height: margin.top + height - yContrib, fill: "#83082A", rx: 2 });
        const subsidyRect = el("rect", { x: bx, y: yTotal, width: barWidth, height: yContrib - yTotal, fill: "#E397A0", rx: 2 });
        const tip = `<strong>${row.age}</strong><br>Prestación contributiva: ${fmt(row.contributiva)}<br>Subsidios: ${fmt(row.subsidios)}<br>Total: ${fmt(row.total)}`;
        [contribRect, subsidyRect].forEach(rect => {
          rect.addEventListener("mousemove", event => showTip(event, tip));
          rect.addEventListener("mouseleave", hideTip);
        });
        el("text", { x: bx + barWidth / 2, y: margin.top + height + 18, "text-anchor": "end", fill: "#404040", "font-size": 11, transform: `rotate(-38 ${bx + barWidth / 2} ${margin.top + height + 18})` }).textContent = row.age;
      });

      el("rect", { x: margin.left, y: 8, width: 12, height: 12, fill: "#83082A" });
      el("text", { x: margin.left + 18, y: 18, fill: "#404040", "font-size": 12 }).textContent = "Prestación contributiva";
      el("rect", { x: margin.left + 182, y: 8, width: 12, height: 12, fill: "#E397A0" });
      el("text", { x: margin.left + 200, y: 18, fill: "#404040", "font-size": 12 }).textContent = "Subsidios de desempleo";
    }
    draw(years[years.length - 1]);
"""


REGIONAL_SUBSIDY_SHARE_JS = r"""
    const years = Object.keys(chartData).map(Number).sort((a, b) => a - b);
    const slider = document.getElementById("yearSlider");
    const yearLabel = document.getElementById("yearLabel");
    const headline = document.getElementById("headline");
    const svg = document.getElementById("peso_subsidios_ccaa");
    const tooltip = document.getElementById("tooltip");
    slider.min = 0;
    slider.max = years.length - 1;
    slider.value = years.length - 1;
    slider.addEventListener("input", () => draw(years[Number(slider.value)]));

    function el(name, attrs = {}, parent = svg) {
      const node = document.createElementNS("http://www.w3.org/2000/svg", name);
      for (const [key, value] of Object.entries(attrs)) node.setAttribute(key, value);
      parent.appendChild(node);
      return node;
    }
    function clear() { while (svg.firstChild) svg.removeChild(svg.firstChild); }
    function fmt(value) { return value.toLocaleString("es-ES", { maximumFractionDigits: 1 }); }
    function showTip(event, html) {
      tooltip.innerHTML = html;
      tooltip.style.left = `${event.clientX}px`;
      tooltip.style.top = `${event.clientY}px`;
      tooltip.style.opacity = 1;
    }
    function hideTip() { tooltip.style.opacity = 0; }
    function draw(year) {
      clear();
      yearLabel.textContent = year;
      const payload = chartData[year];
      const rows = [...payload.regions].sort((a, b) => a.share - b.share);
      const spain = payload.spain;
      headline.textContent = `${fmt(spain)} España`;
      const margin = { top: 18, right: 36, bottom: 42, left: 190 };
      const width = 920 - margin.left - margin.right;
      const height = 430 - margin.top - margin.bottom;
      const maxValue = 100;
      const step = height / rows.length;
      const x = value => margin.left + value / maxValue * width;
      const y = index => margin.top + index * step + step * 0.16;

      for (let tick = 0; tick <= 100; tick += 20) {
        const tx = x(tick);
        el("line", { x1: tx, y1: margin.top, x2: tx, y2: margin.top + height, stroke: "#D9D9D9" });
        el("text", { x: tx, y: margin.top + height + 22, "text-anchor": "middle", fill: "#404040", "font-size": 12 }).textContent = tick;
      }
      const sx = x(spain);
      el("line", { x1: sx, y1: margin.top - 4, x2: sx, y2: margin.top + height, stroke: "#404040", "stroke-width": 3 });
      el("text", { x: sx + 6, y: margin.top + 12, fill: "#404040", "font-size": 12, "font-weight": 700 }).textContent = "España";

      rows.forEach((row, i) => {
        const barWidth = Math.max(0, x(row.share) - margin.left);
        el("text", { x: margin.left - 10, y: y(i) + step * 0.35, "text-anchor": "end", fill: "#404040", "font-size": 12 }).textContent = row.short;
        const rect = el("rect", {
          x: margin.left,
          y: y(i),
          width: barWidth,
          height: step * 0.58,
          fill: row.share >= spain ? "#83082A" : "#E397A0",
          rx: 2
        });
        rect.addEventListener("mousemove", event => showTip(event, `<strong>${row.region}</strong><br>Subsidios: ${fmt(row.share)}<br>España: ${fmt(spain)}`));
        rect.addEventListener("mouseleave", hideTip);
        el("text", { x: x(row.share) + 6, y: y(i) + step * 0.38, fill: "#404040", "font-size": 11 }).textContent = fmt(row.share);
      });
    }
    draw(years[years.length - 1]);
"""


AGE_GENDER_SHARE_JS = r"""
    const years = Object.keys(chartData).map(Number).sort((a, b) => a - b);
    const slider = document.getElementById("yearSlider");
    const yearLabel = document.getElementById("yearLabel");
    const headline = document.getElementById("headline");
    const svg = document.getElementById("peso_mujeres_edad");
    const tooltip = document.getElementById("tooltip");
    slider.min = 0;
    slider.max = years.length - 1;
    slider.value = years.length - 1;
    slider.addEventListener("input", () => draw(years[Number(slider.value)]));

    function el(name, attrs = {}, parent = svg) {
      const node = document.createElementNS("http://www.w3.org/2000/svg", name);
      for (const [key, value] of Object.entries(attrs)) node.setAttribute(key, value);
      parent.appendChild(node);
      return node;
    }
    function clear() { while (svg.firstChild) svg.removeChild(svg.firstChild); }
    function fmt(value) { return value.toLocaleString("es-ES", { maximumFractionDigits: 1 }); }
    function showTip(event, html) {
      tooltip.innerHTML = html;
      tooltip.style.left = `${event.clientX}px`;
      tooltip.style.top = `${event.clientY}px`;
      tooltip.style.opacity = 1;
    }
    function hideTip() { tooltip.style.opacity = 0; }
    function draw(year) {
      clear();
      yearLabel.textContent = year;
      const rows = chartData[year];
      const average = rows.reduce((sum, row) => sum + row.share, 0) / rows.length;
      headline.textContent = `${fmt(average)} media`;
      const margin = { top: 20, right: 26, bottom: 86, left: 64 };
      const width = 920 - margin.left - margin.right;
      const height = 430 - margin.top - margin.bottom;
      const maxValue = 100;
      const band = width / rows.length;
      const x = index => margin.left + index * band + band * 0.18;
      const barWidth = band * 0.64;
      const y = value => margin.top + height - value / maxValue * height;

      for (let tick = 0; tick <= maxValue; tick += 20) {
        const ty = y(tick);
        el("line", { x1: margin.left, y1: ty, x2: margin.left + width, y2: ty, stroke: "#D9D9D9" });
        el("text", { x: margin.left - 10, y: ty + 4, "text-anchor": "end", fill: "#404040", "font-size": 12 }).textContent = tick;
      }
      el("line", { x1: margin.left, y1: y(50), x2: margin.left + width, y2: y(50), stroke: "#404040", "stroke-width": 2 });
      rows.forEach((row, i) => {
        const bx = x(i);
        const by = y(row.share);
        const rect = el("rect", {
          x: bx,
          y: by,
          width: barWidth,
          height: margin.top + height - by,
          fill: row.share >= 50 ? "#83082A" : "#E397A0",
          rx: 2
        });
        rect.addEventListener("mousemove", event => showTip(event, `<strong>${row.age}</strong><br>Mujeres: ${fmt(row.share)}<br>Hombres: ${fmt(100 - row.share)}`));
        rect.addEventListener("mouseleave", hideTip);
        el("text", { x: bx + barWidth / 2, y: margin.top + height + 18, "text-anchor": "end", fill: "#404040", "font-size": 11, transform: `rotate(-38 ${bx + barWidth / 2} ${margin.top + height + 18})` }).textContent = row.age;
      });
    }
    draw(years[years.length - 1]);
"""


COVERAGE_HEATMAP_JS = r"""
    document.querySelector(".controls").style.display = "none";
    const years = Object.keys(chartData).map(Number).sort((a, b) => a - b);
    const svg = document.getElementById("mapa_calor_cobertura");
    const tooltip = document.getElementById("tooltip");
    const headline = document.getElementById("headline");
    headline.textContent = `${years[0]}-${years[years.length - 1]}`;

    function el(name, attrs = {}, parent = svg) {
      const node = document.createElementNS("http://www.w3.org/2000/svg", name);
      for (const [key, value] of Object.entries(attrs)) node.setAttribute(key, value);
      parent.appendChild(node);
      return node;
    }
    function fmt(value) { return value.toLocaleString("es-ES", { maximumFractionDigits: 1 }); }
    function showTip(event, html) {
      tooltip.innerHTML = html;
      tooltip.style.left = `${event.clientX}px`;
      tooltip.style.top = `${event.clientY}px`;
      tooltip.style.opacity = 1;
    }
    function hideTip() { tooltip.style.opacity = 0; }
    const regions = chartData[years[0]].regions.map(row => row.region);
    const allValues = years.flatMap(year => chartData[year].regions.map(row => row.coverage));
    const minValue = Math.floor(Math.min(...allValues) / 5) * 5;
    const maxValue = Math.ceil(Math.max(...allValues) / 5) * 5;
    const margin = { top: 26, right: 28, bottom: 48, left: 190 };
    const width = 920 - margin.left - margin.right;
    const height = 430 - margin.top - margin.bottom;
    const cellW = width / years.length;
    const cellH = height / regions.length;
    function color(value) {
      const t = Math.max(0, Math.min(1, (value - minValue) / (maxValue - minValue)));
      const start = [243, 214, 218];
      const end = [131, 8, 42];
      const rgb = start.map((s, i) => Math.round(s + (end[i] - s) * t));
      return `rgb(${rgb[0]},${rgb[1]},${rgb[2]})`;
    }
    years.forEach((year, xIndex) => {
      el("text", { x: margin.left + xIndex * cellW + cellW / 2, y: margin.top + height + 22, "text-anchor": "middle", fill: "#404040", "font-size": 12 }).textContent = year;
      const byRegion = Object.fromEntries(chartData[year].regions.map(row => [row.region, row.coverage]));
      regions.forEach((region, yIndex) => {
        const value = byRegion[region];
        const rect = el("rect", {
          x: margin.left + xIndex * cellW,
          y: margin.top + yIndex * cellH,
          width: cellW - 1,
          height: cellH - 1,
          fill: color(value)
        });
        rect.addEventListener("mousemove", event => showTip(event, `<strong>${region}</strong><br>${year}: ${fmt(value)}`));
        rect.addEventListener("mouseleave", hideTip);
      });
    });
    regions.forEach((region, yIndex) => {
      el("text", { x: margin.left - 10, y: margin.top + yIndex * cellH + cellH * 0.66, "text-anchor": "end", fill: "#404040", "font-size": 12 }).textContent = chartData[years[0]].regions.find(row => row.region === region).short;
    });
    el("text", { x: margin.left, y: 16, fill: "#404040", "font-size": 12 }).textContent = `Escala fija: ${minValue}-${maxValue}`;
"""


def plot_regional_dispersion(csv_path: Path) -> Path:
    rows = read_regional_dispersion(csv_path)
    if not rows:
        return Path()
    periods = [row["period"] for row in rows]
    minimum = [row["min"] for row in rows]
    maximum = [row["max"] for row in rows]
    median = [row["median"] for row in rows]
    national = read_national_all_ages(csv_path)
    national_by_period = {
        row["period"]: safe_sum(row.get("total prestacion contributiva"), row.get("total subsidios de desempleo"))
        for row in national
    }
    national_index = make_index([national_by_period.get(period) for period in periods])

    fig, ax = plt.subplots(figsize=(6, 3))
    ax.fill_between(periods, minimum, maximum, color="#E397A0", alpha=0.45, label="Rango CCAA")
    ax.plot(periods, median, color="#83082A", linewidth=2, label="Mediana CCAA")
    ax.plot(periods, national_index, color="#404040", linewidth=1.5, linestyle="--", label="España")
    ax.axhline(100, color="#D9D9D9", linewidth=1, linestyle=":")
    ax.xaxis.set_major_locator(mdates.YearLocator())
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
    ax.grid(axis="y")
    ax.grid(axis="x", visible=False)
    ax.tick_params(direction="out", colors="#404040")
    ax.legend(loc="upper left", frameon=False, fontsize=8)
    write_chart_workbook(
        "dispersion_ccaa_beneficiarios.xlsx",
        "Dispersión territorial de los beneficiarios",
        ["Periodo", "Mínimo CCAA", "Mediana CCAA", "Máximo CCAA", "España"],
        zip(periods, minimum, median, maximum, national_index),
        "Fuente: AIReF a partir de SEPE.",
        "Nota: Índice base primer mes = 100. Beneficiarios con prestación contributiva o subsidio de desempleo.",
    )
    return save_figure(fig, "dispersion_ccaa_beneficiarios.svg")


def read_age_profile(csv_path: Path) -> list[dict]:
    rows = []
    latest = latest_period(csv_path)
    age_order = {
        "16 - 19 años": 1,
        "20 - 24 años": 2,
        "25 - 29 años": 3,
        "30 - 34 años": 4,
        "35 - 39 años": 5,
        "40 - 44 años": 6,
        "45 - 49 años": 7,
        "50 - 54 años": 8,
        "55 - 59 años": 9,
        "60 y más años": 10,
    }
    with csv_path.open(newline="", encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            period = date(int(row["año"]), int(row["mes"]), 1)
            if period != latest:
                continue
            if row.get("sexo") != "Ambos sexos" or row.get("nivel geografico") != "espana":
                continue
            age = row.get("edad")
            if age not in age_order:
                continue
            contributiva = parse_csv_number(row.get("total prestacion contributiva")) or 0
            subsidios = parse_csv_number(row.get("total subsidios de desempleo")) or 0
            rows.append({
                "age": age,
                "contributiva": contributiva,
                "subsidios": subsidios,
                "total": contributiva + subsidios,
                "period_label": period_label(period),
            })
    return sorted(rows, key=lambda row: age_order[row["age"]])


def read_gender_share(csv_path: Path) -> list[dict]:
    by_period: dict[date, dict[str, float]] = {}
    with csv_path.open(newline="", encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            if row.get("edad") != "Todas las edades" or row.get("nivel geografico") != "espana":
                continue
            if row.get("sexo") not in {"Hombres", "Mujeres"}:
                continue
            period = date(int(row["año"]), int(row["mes"]), 1)
            total = safe_sum(
                parse_csv_number(row.get("total prestacion contributiva")),
                parse_csv_number(row.get("total subsidios de desempleo")),
            )
            if total is not None:
                by_period.setdefault(period, {})[row["sexo"]] = total
    rows = []
    for period, values in sorted(by_period.items()):
        men = values.get("Hombres")
        women = values.get("Mujeres")
        if men and women:
            rows.append({"period": period, "women_share": women / (men + women) * 100})
    return rows


def read_regional_latest(csv_path: Path) -> list[dict]:
    latest = latest_period(csv_path)
    rows = []
    with csv_path.open(newline="", encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            period = date(int(row["año"]), int(row["mes"]), 1)
            if period != latest:
                continue
            if row.get("sexo") != "Ambos sexos" or row.get("edad") != "Todas las edades":
                continue
            if row.get("nivel geografico") != "comunidad_autonoma":
                continue
            rows.append({
                "region": row.get("comunidad autonoma"),
                "coverage": parse_csv_number(row.get("tasa de cobertura")),
                "period_label": period_label(period),
            })
    return rows


def read_latest_national_coverage(csv_path: Path):
    latest = latest_period(csv_path)
    with csv_path.open(newline="", encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            period = date(int(row["año"]), int(row["mes"]), 1)
            if period != latest:
                continue
            if row.get("sexo") != "Ambos sexos" or row.get("edad") != "Todas las edades":
                continue
            if row.get("nivel geografico") == "espana":
                return parse_csv_number(row.get("tasa de cobertura"))
    return None


def read_regional_coverage_by_year(csv_path: Path) -> dict[int, dict]:
    regions: dict[int, dict[str, list[float]]] = {}
    spain: dict[int, list[float]] = {}
    with csv_path.open(newline="", encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            if row.get("sexo") != "Ambos sexos" or row.get("edad") != "Todas las edades":
                continue
            coverage = parse_csv_number(row.get("tasa de cobertura"))
            if coverage is None:
                continue
            year = int(row["año"])
            level = row.get("nivel geografico")
            if level == "comunidad_autonoma":
                region = row.get("comunidad autonoma")
                regions.setdefault(year, {}).setdefault(region, []).append(coverage)
            elif level == "espana":
                spain.setdefault(year, []).append(coverage)

    out = {}
    for year in sorted(regions):
        if year not in spain:
            continue
        out[year] = {
            "spain": average(spain[year]),
            "regions": [
                {"region": region, "short": short_region_name(region), "coverage": average(values)}
                for region, values in sorted(regions[year].items())
            ],
        }
    return out


def read_age_profile_by_year(csv_path: Path) -> dict[int, list[dict]]:
    age_order = {
        "16 - 19 años": 1,
        "20 - 24 años": 2,
        "25 - 29 años": 3,
        "30 - 34 años": 4,
        "35 - 39 años": 5,
        "40 - 44 años": 6,
        "45 - 49 años": 7,
        "50 - 54 años": 8,
        "55 - 59 años": 9,
        "60 y más años": 10,
    }
    grouped: dict[int, dict[str, dict[str, list[float]]]] = {}
    with csv_path.open(newline="", encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            if row.get("sexo") != "Ambos sexos" or row.get("nivel geografico") != "espana":
                continue
            age = row.get("edad")
            if age not in age_order:
                continue
            year = int(row["año"])
            bucket = grouped.setdefault(year, {}).setdefault(age, {"contributiva": [], "subsidios": []})
            bucket["contributiva"].append(parse_csv_number(row.get("total prestacion contributiva")) or 0)
            bucket["subsidios"].append(parse_csv_number(row.get("total subsidios de desempleo")) or 0)

    out = {}
    for year in sorted(grouped):
        rows = []
        for age, values in grouped[year].items():
            contributiva = average(values["contributiva"])
            subsidios = average(values["subsidios"])
            rows.append({
                "age": age,
                "contributiva": contributiva,
                "subsidios": subsidios,
                "total": contributiva + subsidios,
            })
        out[year] = sorted(rows, key=lambda row: age_order[row["age"]])
    return out


def read_regional_subsidy_share_by_year(csv_path: Path) -> dict[int, dict]:
    regions: dict[int, dict[str, dict[str, float]]] = {}
    spain: dict[int, dict[str, float]] = {}
    with csv_path.open(newline="", encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            if row.get("sexo") != "Ambos sexos" or row.get("edad") != "Todas las edades":
                continue
            contributiva = parse_csv_number(row.get("total prestacion contributiva")) or 0
            subsidios = parse_csv_number(row.get("total subsidios de desempleo")) or 0
            if contributiva + subsidios == 0:
                continue
            year = int(row["año"])
            level = row.get("nivel geografico")
            if level == "comunidad_autonoma":
                region = row.get("comunidad autonoma")
                bucket = regions.setdefault(year, {}).setdefault(region, {"contributiva": 0, "subsidios": 0})
            elif level == "espana":
                bucket = spain.setdefault(year, {"contributiva": 0, "subsidios": 0})
            else:
                continue
            bucket["contributiva"] += contributiva
            bucket["subsidios"] += subsidios

    out = {}
    for year in sorted(regions):
        if year not in spain:
            continue
        spain_total = spain[year]["contributiva"] + spain[year]["subsidios"]
        out[year] = {
            "spain": spain[year]["subsidios"] / spain_total * 100,
            "regions": [
                {
                    "region": region,
                    "short": short_region_name(region),
                    "share": values["subsidios"] / (values["contributiva"] + values["subsidios"]) * 100,
                }
                for region, values in sorted(regions[year].items())
                if values["contributiva"] + values["subsidios"]
            ],
        }
    return out


def read_age_gender_share_by_year(csv_path: Path) -> dict[int, list[dict]]:
    age_order = {
        "16 - 19 años": 1,
        "20 - 24 años": 2,
        "25 - 29 años": 3,
        "30 - 34 años": 4,
        "35 - 39 años": 5,
        "40 - 44 años": 6,
        "45 - 49 años": 7,
        "50 - 54 años": 8,
        "55 - 59 años": 9,
        "60 y más años": 10,
    }
    grouped: dict[int, dict[str, dict[str, float]]] = {}
    with csv_path.open(newline="", encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            if row.get("nivel geografico") != "espana" or row.get("sexo") not in {"Hombres", "Mujeres"}:
                continue
            age = row.get("edad")
            if age not in age_order:
                continue
            total = safe_sum(
                parse_csv_number(row.get("total prestacion contributiva")),
                parse_csv_number(row.get("total subsidios de desempleo")),
            )
            if total is None:
                continue
            year = int(row["año"])
            grouped.setdefault(year, {}).setdefault(age, {"Hombres": 0, "Mujeres": 0})[row["sexo"]] += total

    out = {}
    for year in sorted(grouped):
        rows = []
        for age, values in grouped[year].items():
            total = values["Hombres"] + values["Mujeres"]
            if total:
                rows.append({"age": age, "share": values["Mujeres"] / total * 100})
        out[year] = sorted(rows, key=lambda row: age_order[row["age"]])
    return out


def read_regional_dispersion(csv_path: Path) -> list[dict]:
    by_period: dict[date, dict[str, float]] = {}
    with csv_path.open(newline="", encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            if row.get("sexo") != "Ambos sexos" or row.get("edad") != "Todas las edades":
                continue
            if row.get("nivel geografico") != "comunidad_autonoma":
                continue
            period = date(int(row["año"]), int(row["mes"]), 1)
            total = safe_sum(
                parse_csv_number(row.get("total prestacion contributiva")),
                parse_csv_number(row.get("total subsidios de desempleo")),
            )
            if total is not None:
                by_period.setdefault(period, {})[row["comunidad autonoma"]] = total
    rows = []
    first_period = min(by_period) if by_period else None
    if not first_period:
        return rows
    regions = sorted(by_period[first_period])
    for period in sorted(by_period):
        values = by_period[period]
        if sorted(values) != regions:
            continue
        period_index = [
            values[region] / by_period[first_period][region] * 100
            for region in regions
            if by_period[first_period][region]
        ]
        if period_index:
            rows.append({
                "period": period,
                "min": min(period_index),
                "median": median(period_index),
                "max": max(period_index),
            })
    return rows


def latest_period(csv_path: Path) -> date:
    latest = None
    with csv_path.open(newline="", encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            period = date(int(row["año"]), int(row["mes"]), 1)
            latest = period if latest is None or period > latest else latest
    if latest is None:
        raise ValueError(f"No periods found in {csv_path}")
    return latest


def make_index(values: list[float | None]) -> list[float | None]:
    base = next((value for value in values if value), None)
    if not base:
        return [None for _ in values]
    return [value / base * 100 if value else None for value in values]


def median(values: list[float]) -> float:
    ordered = sorted(values)
    midpoint = len(ordered) // 2
    if len(ordered) % 2:
        return ordered[midpoint]
    return (ordered[midpoint - 1] + ordered[midpoint]) / 2


def average(values: list[float]) -> float:
    return sum(values) / len(values) if values else 0


def period_label(period: date) -> str:
    return f"{period.year}-{period.month:02d}"


def short_region_name(name: str) -> str:
    replacements = {
        "Asturias, Principado de": "Asturias",
        "Madrid, Comunidad de": "Madrid",
        "Murcia, Región de": "Murcia",
        "Navarra, Comunidad Foral de": "Navarra",
    }
    return replacements.get(name, name)


def save_figure(fig, filename: str) -> Path:
    path = FIGURES_DIR / filename
    fig.tight_layout()
    fig.savefig(path, bbox_inches="tight")
    if path.suffix.lower() == ".svg":
        fig.savefig(path.with_suffix(".png"), dpi=300, bbox_inches="tight")
    plt.close(fig)
    return path


def write_chart_workbook(filename: str, title: str, headers: list[str], rows, source: str, note: str) -> Path:
    FIGURE_WORKBOOKS_DIR.mkdir(parents=True, exist_ok=True)
    path = FIGURE_WORKBOOKS_DIR / filename
    wb = Workbook()
    ws = wb.active
    ws.title = "grafico"
    ws["B2"] = title
    ws["B3"] = source
    ws["B4"] = note
    for col, header in enumerate(headers, start=4):
        ws.cell(row=5, column=col, value=header)
    for row_idx, values in enumerate(rows, start=6):
        for col_idx, value in enumerate(values, start=4):
            ws.cell(row=row_idx, column=col_idx, value=value)
    for col in ws.columns:
        header = str(col[0].value or "")
        ws.column_dimensions[col[0].column_letter].width = min(max(len(header) + 2, 12), 35)
    wb.save(path)
    return path


def safe_sum(left, right):
    values = [value for value in (left, right) if isinstance(value, (int, float))]
    return sum(values) if values else None


def format_thousands(value, _position):
    return f"{value / 1_000_000:.1f}M" if abs(value) >= 1_000_000 else f"{value:,.0f}".replace(",", ".")


def format_percent_no_symbol(value, _position):
    return f"{value:.0f}"


def make_wide(records: list[dict], all_ages_only: bool) -> list[dict]:
    grouped: dict[tuple, dict] = {}
    for rec in records:
        if all_ages_only and rec["edad"] != "Todas las edades":
            continue
        key = tuple(rec[k] for k in WIDE_KEY_FIELDS)
        row = grouped.setdefault(key, {k: rec[k] for k in WIDE_KEY_FIELDS})
        col = rec["variable"]
        value = rec["valor"]
        if col not in row or row[col] in (None, ""):
            row[col] = value
    for row in grouped.values():
        fill_no_cotizacion_aggregate(row)
        ordered = {k: row.get(k) for k in WIDE_KEY_FIELDS}
        for field in WIDE_VALUE_FIELDS:
            if field in row:
                ordered[field] = row[field]
        row.clear()
        row.update(ordered)
    return sorted(grouped.values(), key=lambda r: (r["año"], r["mes"], r["sexo"], r["nivel geografico"], r["comunidad autonoma"], r["provincia"], r["edad"]))


def fill_no_cotizacion_aggregate(row: dict) -> None:
    aggregate = "subsidio de desempleo por no cotizacion suficiente"
    parts = [
        "subsidio de desempleo por no cotizacion suficiente - derecho de 3 a 5 meses",
        "subsidio de desempleo por no cotizacion suficiente - derecho de 6 meses",
        "subsidio de desempleo por no cotizacion suficiente - derecho de 21 meses",
    ]
    values = [row.get(part) for part in parts]
    numeric = [value for value in values if isinstance(value, (int, float))]
    if numeric:
        row[aggregate] = sum(numeric)


def slug(text: str) -> str:
    replacements = str.maketrans("áéíóúüñÁÉÍÓÚÜÑ", "aeiouunAEIOUUN")
    text = norm_text(text).translate(replacements).lower()
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    return text or "value"


def write_xlsx(rows: list[dict], path: Path, headers: list[str] | None = None) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "sepe_prestaciones_wide"
    headers = headers or list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])
    ws.freeze_panes = "A2"
    for col in ws.columns:
        header = str(col[0].value or "")
        ws.column_dimensions[col[0].column_letter].width = min(max(len(header) + 2, 12), 45)
    wb.save(path)


def process_paths(paths: list[Path]) -> list[dict]:
    manifest = load_manifest()
    by_name = {Path(v.get("local_path", "")).name: v.get("source_url", "") for v in manifest.get("files", {}).values()}
    records = []
    for path in paths:
        records.extend(parse_workbook(path, by_name.get(path.name)))
    export_records(records)
    return records


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Download and reshape SEPE prestaciones Excel workbooks.")
    parser.add_argument("--no-download", action="store_true", help="Process existing files in data/raw only.")
    parser.add_argument("--limit", type=int, help="Limit discovered downloads for smoke tests.")
    parser.add_argument("--from-year", type=int, help="Only include files from this year onward.")
    parser.add_argument("--to-year", type=int, help="Only include files up to this year.")
    args = parser.parse_args(argv)

    if args.no_download:
        paths = sorted(RAW_DIR.glob("*.xlsx"))
    else:
        files = discover_files()
        if args.from_year:
            files = [f for f in files if f.year >= args.from_year]
        if args.to_year:
            files = [f for f in files if f.year <= args.to_year]
        paths = download_files(files, args.limit)
    records = process_paths(paths)
    print(f"Processed {len(paths)} workbook(s), {len(records)} long records.")
    print(f"Wrote {PROCESSED_DIR / 'sepe_prestaciones_wide.csv'}")


if __name__ == "__main__":
    main(sys.argv[1:])
